import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter.messagebox import showinfo 
from tkinter import filedialog
import pyperclip
import os
import pandas as pd
import re
from tkinter import *
from tkinter.ttk import *
import threading
import queue
from functools import partial
import time
import tkinter.scrolledtext as st 
from numpy import interp
import xlrd 
import cantools
from pprint import pprint
import can
from openpyxl.styles import Alignment
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side


window = tk.Tk()
window.title("Language Validation Tool")
window.geometry('1000x530')
window .iconbitmap(r"fca.ico")
window.configure(background='teal')
par=ttk.Notebook(window)
tab_1=ttk.Frame(par)
tab_2=ttk.Frame(par)
tab_3=ttk.Frame(par)
par.add(tab_1,text="WAT Viewer")
par.add(tab_2,text="Test case verify")
par.add(tab_3,text="Template generator")
par.pack(expand=1,fill='both')
#par.configure(background='teal')
def clicked():
    print('WORKING')
    print(langchoosen.get())
    print(pyperclip.paste())
    
def clicked1():
    print('WORKING')
    line1.config(state='normal')
    
    
def browse():
    txt1.delete('0', END)
    filename=filedialog.askopenfilename(initialdir = os.path.expanduser('~')+'\Desktop',title = "Select WAT DB",filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
    txt1.insert(END,filename)
    df = pd.read_excel(filename, index_col=0)
    tmp=df.columns.values.tolist()
    if 'Common ID' and 'NAFTA ID' in tmp:
        startfrom=tmp.index('US English')
        endof=tmp.index('Translation Context')
        global pycommoniddf
        global pynaftaiddf
        pycommoniddf = df['Common ID'].tolist()
        pynaftaiddf =df['NAFTA ID'].tolist()
        global languages
        languages={}
        dowpdownlanguage=[]
        idflag=1
        for y in range(startfrom,endof):
            exec('global py'+str(re.sub(r'[^A-Za-z]','', tmp[y]))+';''py'+str(re.sub(r'[^A-Za-z]','', tmp[y]))+'='+'df[\''+tmp[y]+'\'].tolist()')
            languages[tmp[y]]='py'+str(re.sub(r'[^A-Za-z]','', tmp[y]))
            dowpdownlanguage.append(tmp[y])
        dowpdownlanguage.pop(0)
        langchoosen['values'] = dowpdownlanguage
        langchoosen.current(0) 
    else:
        messagebox.showerror("Error", "Load WAT DB") 
        
    
btn1 = tk.Button(tab_1, height=1, width=7,text="Load DB",bg = "cadetblue4", command=browse)
btn1.place(x=380, y=22)
txt1 = tk.Entry(tab_1, width=60)
txt1.place(x=10, y=25)
l1=tk.Label(tab_1,height=1, width=10,text="WAT DB Path :")
l1.place(x=10, y=1)
l2=tk.Label(tab_1,height=1, width=10,text="WAT Browser :")
l2.place(x=10, y=80)
l3=tk.Label(tab_1,height=1, width=10,text="Search by :")
l3.place(x=10, y=105)
n = tk.StringVar() 
searchchoosen = ttk.Combobox(tab_1, width = 12, textvariable = n,justify='center',state="readonly") 
# Adding combobox drop down list 
searchchoosen['values'] = ('Common ID','NAFTA ID') 
searchchoosen.place(x=85, y=105) 
searchchoosen.current(0) 
l4=tk.Label(tab_1,height=1, width=5,text="ID :")
l4.place(x=15, y=150)
txt2 = tk.Entry(tab_1, width=15)
txt2.place(x=50, y=151)

def find():
    watid=txt2.get().strip()
    if len(watid)==0:
        messagebox.showwarning("Warning", "Empty Field in ID") 
    else:
        if watid.endswith('_1') or watid.endswith('_2') or watid.endswith('_3') or watid.endswith('_4') or watid.endswith('_5') or watid.endswith('_6'):
            watid=watid[:-2]
        if watid.endswith('_'):
            watid=watid[:-1]
        watid_1=watid+'_1'
        if searchchoosen.get().strip() == "Common ID":
            idflag=0
        else:
            idflag=1
        selectlanguage=langchoosen.get().strip()
        try:
            if idflag == 0:
                rowindex=pycommoniddf.index(watid_1)
            else:
                rowindex=pynaftaiddf.index(watid_1)
            global englisharray
            global languagearray
            englisharray=[]
            languagearray=[]
            line=1
            for i in range(rowindex,rowindex+6):
                eval("""line"""+str(line)+""".config(state='normal')""")
                eval("""line"""+str(line)+""".delete('0', END)""")
                if watid in pycommoniddf[i] or watid in pynaftaiddf[i]:
                    if len(str(pyUSEnglish[i])) == 3 and str(pyUSEnglish[i]) == 'nan':
                        englisharray.append('')
                        eval("line"+str(line)+".insert(END,\"\")")
                    else:
                        englisharray.append(pyUSEnglish[i])
                        eval('line'+str(line)+'.insert(END,\"'+pyUSEnglish[i]+'\")')
                eval("""line"""+str(line)+""".config(state='readonly')""")
                if line < 7:
                    line=line+1
            line=1
            for i in range(rowindex,rowindex+6):
                eval("""line"""+str(line)+"""a.config(state='normal')""")
                eval("""line"""+str(line)+"""a.delete('0', END)""")
                if watid in pycommoniddf[i] or watid in pynaftaiddf[i]:
                    if len(str(eval(languages[selectlanguage]+'['+str(i)+']'))) == 3 and str(eval(languages[selectlanguage]+'['+str(i)+']')) == 'nan':
                        languagearray.append('')
                        eval("line"+str(line)+"a.insert(END,\"\")")
                    else:
                        languagearray.append(eval(languages[selectlanguage]+'['+str(i)+']'))
                        eval('line'+str(line)+'a.insert(END,\"'+eval(languages[selectlanguage]+'['+str(i)+']')+'\")')
                eval("""line"""+str(line)+"""a.config(state='readonly')""")
                if line < 7:
                    line=line+1
        except:
            messagebox.showwarning("Warning", "WAT ID not found") 


btn3 = tk.Button(tab_1, height=1, width=7,text="Find",bg = "cadetblue4", command=find)
btn3.place(x=150, y=148)

def pasteandsearch():
    fromclip=pyperclip.paste()
    if len(fromclip) > 2:
        txt2.delete('0', END)
        txt2.insert(END,fromclip)
        find()
    else:
        messagebox.showinfo("info", "Clipboard doesn't have ID") 
    

btn4 = tk.Button(tab_1, height=1, width=10,text="Paste & Find",bg = "cadetblue4", command=pasteandsearch)
btn4.place(x=220, y=148)
l3=tk.Label(tab_1,height=1, width=18,text="Language : US English")
l3.place(x=100, y=200)
l4=tk.Label(tab_1,height=1, width=7,text="Line 1 :")
l4.place(x=10, y=250)
l5=tk.Label(tab_1,height=1, width=7,text="Line 2 :")
l5.place(x=10, y=280)
l6=tk.Label(tab_1,height=1, width=7,text="Line 3 :")
l6.place(x=10, y=310)
l7=tk.Label(tab_1,height=1, width=7,text="Line 4 :")
l7.place(x=10, y=340)
l8=tk.Label(tab_1,height=1, width=7,text="Line 5 :")
l8.place(x=10, y=370)
l9=tk.Label(tab_1,height=1, width=7,text="Line 6 :")
l9.place(x=10, y=400)
line1 = tk.Entry(tab_1, width=35,justify='center')
line1.place(x=70, y=251)
line2 = tk.Entry(tab_1, width=35,justify='center')
line2.place(x=70, y=281)
line3 = tk.Entry(tab_1, width=35,justify='center')
line3.place(x=70, y=311)
line4 = tk.Entry(tab_1, width=35,justify='center')
line4.place(x=70, y=341)
line5 = tk.Entry(tab_1, width=35,justify='center')
line5.place(x=70, y=371)
line6 = tk.Entry(tab_1, width=35,justify='center')
line6.place(x=70, y=401)

def copylines():
    str1='\n'
    try:
        pyperclip.copy(str1.join(englisharray))
    except:
        messagebox.showwarning("Warning", "Empty field in message lines") 

def copyline():
    str1=' '
    try:
        pyperclip.copy(str1.join(englisharray))
    except:
        messagebox.showwarning("Warning", "Empty field in message lines")


btn5 = tk.Button(tab_1, height=1, width=10,text="Copy Lines",bg = "cadetblue4", command=copylines)
btn5.place(x=80, y=435)
btn6 = tk.Button(tab_1, height=1, width=11,text="Copy as a Line",bg = "cadetblue4", command=copyline)
btn6.place(x=185, y=435)
l3=tk.Label(tab_1,height=1, width=10,text="Language :")
l3.place(x=470, y=200)
m = tk.StringVar() 
langchoosen = ttk.Combobox(tab_1,width = 15,textvariable = m,justify='center',state="readonly") 
langchoosen.place(x=540, y=201) 
l10=tk.Label(tab_1,height=1,width=7,text="Line 1 :")
l10.place(x=400, y=250)
l11=tk.Label(tab_1,height=1,width=7,text="Line 2 :")
l11.place(x=400, y=280)
l12=tk.Label(tab_1,height=1,width=7,text="Line 3 :")
l12.place(x=400, y=310)
l13=tk.Label(tab_1,height=1,width=7,text="Line 4 :")
l13.place(x=400, y=340)
l14=tk.Label(tab_1,height=1,width=7,text="Line 5 :")
l14.place(x=400, y=370)
l15=tk.Label(tab_1,height=1,width=7,text="Line 6 :")
l15.place(x=400, y=400)
line1a = tk.Entry(tab_1,width=35,justify='center')
line1a.place(x=460, y=251)
line2a = tk.Entry(tab_1,width=35,justify='center')
line2a.place(x=460, y=281)
line3a = tk.Entry(tab_1,width=35,justify='center')
line3a.place(x=460, y=311)
line4a = tk.Entry(tab_1,width=35,justify='center')
line4a.place(x=460, y=341)
line5a = tk.Entry(tab_1,width=35,justify='center')
line5a.place(x=460, y=371)
line6a = tk.Entry(tab_1,width=35,justify='center')
line6a.place(x=460, y=401)

def copylineslang():
    str1='\n'
    try:
        pyperclip.copy(str1.join(languagearray))
    except:
        messagebox.showwarning("Warning", "Empty field in message lines") 

def copylinelang():
    str1=' '
    try:
        pyperclip.copy(str1.join(languagearray))
    except:
        messagebox.showwarning("Warning","Empty field in message lines") 

btn5a = tk.Button(tab_1, height=1, width=10,text="Copy Lines",bg = "cadetblue4", command=copylineslang)
btn5a.place(x=480, y=435)
btn6a = tk.Button(tab_1, height=1, width=11,text="Copy as a Line",bg = "cadetblue4", command=copylinelang)
btn6a.place(x=585, y=435)





progress=Progressbar(tab_2,orient=HORIZONTAL,length=800,mode='determinate')
progress.place(x=50, y=400)


def worker(q, r,update):
        #Passing the current value to the queue
    q.put(update + 1)
    #Generate event
    r.event_generate('<<Updated>>', when='tail')
    #Sleep for clarity




def on_update(event, q=None, pb=None):
    pb['value']=q.get()
    


q=queue.Queue()  
handler=partial(on_update, q=q, pb=progress)
window.bind('<<Updated>>', handler)

def progress():
    worker(q,window,-1)
    try:
        if len(filename_2) <5:
            tk.messagebox.showinfo('Error','Load Testcase')
        elif len(filename2_2) <5:
            tk.messagebox.showinfo('Error','Load DBC')
    except:
        tk.messagebox.showinfo('Error','Load Testcase & DBC')
    try:
        global wb
        global dbct
        global sheet
        wb = xlrd.open_workbook(filename_2) 
        sheet = wb.sheet_by_name('Indication') 
        sheet.cell_value(0, 0) 
        dbct = cantools.database.load_file(filename2_2)
    except:
        tk.messagebox.showinfo('Error','Testcase doest have "Indication" sheet')
    global bus_s
    bus_s=bus_string.get().strip()
    if len(bus_s) < 3 :
        tk.messagebox.showinfo('Error','Enter valid BUS name ')
    else:
        MsgBox = tk.messagebox.askquestion ('Validate TS','Ensure BUS name is : '+bus_s+' \nStart Validate ?',icon = 'warning')
        if MsgBox == 'yes':
           print('Run test')
           t1=threading.Thread(target=testts)
           t1.start()
           tx=threading.Thread(target=console, args=("Test cases"))
           tx.start()
        else:
            tk.messagebox.showinfo('Return','You will now return to the Validate screen')


def console(text):
    print('Working--',text)
    text_area.insert(END,text) 

    


def testts():
    t=threading.Thread(target=worker, args=(q, window,0))
    t.start()
    update=50
    wb = openpyxl.Workbook()
    HMI = wb['Sheet']
    HMI.title = 'Report'
    HMI.freeze_panes =HMI['B2']
    
    blueFill = PatternFill(start_color='68b7f3',end_color='68b7f3',fill_type='solid')
    redfill = PatternFill(start_color='f04242',end_color='f04242',fill_type='solid')
    yellowfill = PatternFill(start_color='dcf155',end_color='dcf155',fill_type='solid')
    greenfill = PatternFill(start_color='5cf042',end_color='5cf042',fill_type='solid')
    
    thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    
    l1=['INDICATION','Signal 1','Signal 2','Signal 3','Signal 4','Signal 5','Signal 6','Signal 7','Signal 8','Signal 9','Signal 10']
    lis=[40,50,50,50,50,50,50,50,50,50,50]
    
    HMI.row_dimensions[1].height= 30
    
    for i in range(0, len(l1)):
        HMI.cell(row=1, column=i+1).value = l1[i]
        HMI.cell(row=1, column=i+1).border = thin_border
        HMI.column_dimensions[str(chr(65+i))].width = lis[i]
        currentCell = HMI.cell(row=1, column=i+1)
        currentCell.alignment = Alignment(horizontal='center',vertical='center')
        HMI.cell(row=1, column=i+1).fill = blueFill

    dbc=open(os.path.join(filename2_2))
    db=list(dbc)
    nol=len(db)
    mslen=len(dbct.messages)
    msgandsignal={}
    bothlist=[]
    print(sheet.nrows)
    for j in range(sheet.nrows):
        if j >= 2:
            HMI.cell(row=j, column=1).value = str(sheet.cell_value(j, 5)).strip()
            HMI.cell(row=j, column=1).border = thin_border
            currentCell = HMI.cell(row=j, column=1) #or currentCell = ws['A1']
            currentCell.alignment = Alignment(horizontal='center',vertical='center')
            
        currentpercentage=int(interp(100,[1,int(sheet.nrows)],[1,int(j)]))
        print(currentpercentage)
        worker(q,window,currentpercentage)
        console(str(sheet.cell_value(j, 5)).strip())
        excelcol=1
        aa=0
        bb=0
        cc=0
        ww=0
        for i in range(sheet.ncols):
            cellvalue=str(sheet.cell_value(j, i)).strip()
            commacount=cellvalue.count(',')
            if commacount > 1 and i > 5 and i < 61:
                excelcol=excelcol+1
                if 'Set CAN Signal' in cellvalue:
                    cellvalue=(cellvalue.split('(')[1].split(')')[0]).strip()
                print(cellvalue)
                signal=cellvalue
                if j>=2:
                    HMI.cell(row=j, column=excelcol).value = str(sheet.cell_value(j, i)).strip()
                    HMI.cell(row=j, column=excelcol).border = thin_border
                    currentCell = HMI.cell(row=j, column=excelcol) #or currentCell = ws['A1']
                    currentCell.alignment = Alignment(horizontal='center',vertical='center')
                if ' ,'  in cellvalue or ', '  in cellvalue:
                    print('/Raise warning Warning/')
                    print(str(sheet.cell_value(j, 5)).strip())
                    text_area.insert(tk.INSERT,str(sheet.cell_value(j, 5)).strip(), 'warning')
                    HMI.cell(row=j, column=excelcol).fill = yellowfill
                    ww=1
                else:           
                    for i in range(0,nol):
                        time_line=str(db[i])
                        if time_line.startswith("BO_"):
                            #print(time_line)
                            tmmsg=(time_line.strip().split(': ')[1].split('\n')[0])
                            msgs=(time_line.strip().split(' ')[2].split(':')[0]).strip()
                            node = (''.join([i for i in tmmsg if not i.isdigit()])).strip()
                            bothlist.append(node+':'+msgs)
                    for ml in range (0,mslen):
                        temp=str(dbct.messages[ml])
                        msglist=((temp.split('(\''))[1].split('\'')[0])
                        msgobjinmsg= dbct.get_message_by_name(msglist)
                        signalcountinmsgobj=len(msgobjinmsg.signals)
                        tempsignallist=[]
                        for sl in range (0,signalcountinmsgobj):
                            signalsinobj=str(msgobjinmsg.signals[sl])
                            sigobj=((signalsinobj.split('(\''))[1].split('\'')[0])
                            tempsignallist.append(sigobj)
                        msgandsignal[msglist]=tempsignallist
                    signalarray=signal.split(',')
                    a=0
                    b=0
                    c=0
                    if signalarray[0] == bus_s:
                        print('Correct bus')
                    else:
                        print('notcorrect bus')
                        a=0
                        aa=1
                 
                    if signalarray[1]+':'+signalarray[2] in bothlist:
                        print('present node and msg')
                    else:
                        print('Not there node or msg')
                        b=1
                        bb=1

                    signalflag=0
                    for key, value in msgandsignal.items():
                        if key == signalarray[2] and signalarray[3] in value:
                            print("Got msg and signal")
                            signalflag=1
                            break
                    if signalflag == 0:
                        print("msg or signal error")
                        c=1
                        cc=1
                    if a==0 and b==0 and c==0:
                        HMI.cell(row=j, column=excelcol).fill = greenfill
                    else:
                        HMI.cell(row=j, column=excelcol).fill = redfill
                        
        if j >= 2:
            if ww==0 and aa==0 and bb==0 and cc==0:
                HMI.cell(row=j, column=1).fill = greenfill
            else:
                HMI.cell(row=j, column=1).fill = redfill



    wb.save(fin+'Report.xlsx')
    tk.messagebox.showinfo('Completed','Report is saved in Testcase path')


global text_area
run = tk.Button(tab_2, height=1, width=11,text="Start",bg = "cadetblue4", command=progress)
run.place(x=880, y=398)
status=tk.Label(tab_2,height=1, width=18,text="Status")
status.place(x=750, y=1)
text_area = st.ScrolledText(tab_2,width=61,height=12,font=("Times New Roman",10)) 
text_area.grid(column = 0, pady = 25, padx = 600) 
text_area.insert(tk.INSERT, 'Set CAN Signal(FDCAN3_1,BSM,BRAKE_FD_2,VehicleSpeedVSOSig,9)')



text_area.insert(tk.INSERT, "Ehila", 'pass')  
text_area.insert(tk.INSERT, "Now", 'error') 
text_area.insert(tk.INSERT, "sig", 'ts')
text_area.insert(tk.INSERT, "war", 'warning')
text_area.tag_config('pass', foreground='green')
text_area.tag_config('error', foreground='red') 
text_area.tag_config('ts', foreground='black') 
text_area.tag_config('warning', foreground='orange') 


text_area.configure(state ='disabled') 



def excelbrowse():
    txt1_2.delete('0', END)
    global filename_2
    global fin
    filename_2=filedialog.askopenfilename(initialdir = os.path.expanduser('~')+'\Desktop',title = "Select Test case",filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
    txt1_2.insert(END,filename_2)
    rem = filename_2.split('/')[-1]
    fin = filename_2.replace(rem,'')



def dbcbrowse():
    global filename2_2
    txt2_2.delete('0', END)
    filename2_2=filedialog.askopenfilename(initialdir = os.path.expanduser('~')+'\Desktop',title = "Select signal DBC",filetypes = (("DBC files","*.dbc"),("all files","*.*")))
    txt2_2.insert(END,filename2_2)


btn1_2 = tk.Button(tab_2, height=1, width=9,text="Load Excel",bg = "cadetblue4", command=excelbrowse)
btn1_2.place(x=380, y=22)
txt1_2 = tk.Entry(tab_2, width=60)
txt1_2.place(x=10, y=25)
l1_2=tk.Label(tab_2,height=1, width=11,text="Test Case Path :")
l1_2.place(x=10, y=1)


btn2_2 = tk.Button(tab_2, height=1, width=9,text="Load DBC",bg = "cadetblue4", command=dbcbrowse)
btn2_2.place(x=380, y=72)
txt2_2 = tk.Entry(tab_2, width=60)
txt2_2.place(x=10, y=75)
l1_3=tk.Label(tab_2,height=1, width=7,text="DBC Path :")
l1_3.place(x=10, y=50)

bus_2=tk.Label(tab_2,height=1, width=8,text="BUS name :")
bus_2.place(x=10, y=120)

bus_string = tk.Entry(tab_2, width=15)
bus_string.place(x=25, y=150)



window.mainloop()